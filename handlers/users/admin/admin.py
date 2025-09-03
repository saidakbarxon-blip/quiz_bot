# handlers/users/admin/admin.py
import os
from datetime import datetime
from aiogram.filters import Command
from aiogram.types import FSInputFile
from keyboards.default.admin_kb import admin_keyboard, channels_button
from utils.database.db import DataBase
import pandas as pd
from openpyxl.styles import Font, PatternFill
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from data.config import load_config
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery
from filters.admin import AdminFilter
from keyboards.inline.channel_actions import get_delete_channel_keyboard


admins: list[int] = load_config().bot.admin_ids

router = Router()
db = DataBase()


# 📌 Kanal qo‘shish va o‘chirish uchun FSM state'lari
class ChannelStates(StatesGroup):
    add_channel = State()
    delete_channel = State()


# 📌 Admin panel
@router.message(AdminFilter(), Command("admin"))
async def admin_panel(message: Message):
    if message.from_user.id not in admins:
        await message.answer("❌ Bu buyruq faqat adminlar uchun!")
        return

    await message.answer(
        "👋 Admin panel:\n\n🔍 Quyidagi funksiyalardan foydalanishingiz mumkin:",
        reply_markup=admin_keyboard,
    )


# 📌 ➕ Kanal qo‘shish
@router.message(AdminFilter(), F.text == "➕ Kanal qo'shish")
async def add_channel(message: Message, state: FSMContext):
    await message.answer(
        "📢 <b>Kanal qo'shish uchun quyidagi formatda ma'lumot kiriting:</b>\n\n"
        "🔹 <code>nom|link|ID</code>\n\n"
        "📝 <b>Misol uchun:</b>\n"
        "<code>myKanal|https://t.me/mykanaluz|-1001234567890</code>\n\n"
        "❗️ <b>Diqqat:</b> Link to'g'ri formatda bo'lishi shart.\n"
        "✅ Faqat <b>https://t.me/</b> bilan boshlanuvchi linklarni kiritish mumkin.\n"
        "🔹 Kanal ID ni olish uchun: <code>@username</code> ni kanalga yuboring va `/id` botidan foydalaning."
    )
    await state.set_state(ChannelStates.add_channel)


@router.message(ChannelStates.add_channel)
async def process_add_channel(message: Message, state: FSMContext):
    try:
        data = message.text.split("|")
        if len(data) != 3:
            raise ValueError(
                "❌ Noto‘g‘ri format! To‘g‘ri format: <code>nom|link|ID</code> bo‘lishi kerak"
            )

        name, link, channel_id = data
        name, link, channel_id = name.strip(), link.strip(), channel_id.strip()

        # Kanal ID raqam ekanligini tekshirish
        if not channel_id.lstrip("-").isdigit():
            raise ValueError(
                "❌ Kanal ID noto‘g‘ri! ID faqat sonlardan iborat bo‘lishi kerak."
            )

        channel_id = int(channel_id)  # Int formatga o‘tkazish

        # Kanalni bazaga qo‘shish
        result = await db.add_subscription(name=name, link=link, channel_id=channel_id)
        await message.answer(f"✅ {result}")  # Bazadan qaytgan natijani yuborish

    except ValueError as e:
        await message.answer(str(e))
    except Exception as e:
        await message.answer(f"❌ Xatolik yuz berdi: {e}")
    finally:
        await state.clear()


# 📌 ➖ Kanal o‘chirish
@router.message(AdminFilter(), F.text == "➖ Kanal o'chirish")
async def delete_channel(message: Message):
    keyboard = await get_delete_channel_keyboard()
    if not keyboard:
        await message.answer("❌ Bazada kanallar mavjud emas!")
        return

    await message.answer(
        "🗑 O'chirmoqchi bo'lgan kanalingizni tanlang:", reply_markup=keyboard
    )


@router.callback_query(F.data.startswith("delete_channel:"))
async def process_delete_channel(callback: CallbackQuery):
    """Tanlangan kanalni bazadan o‘chirish."""
    subscription_id = int(callback.data.split(":")[1])

    try:
        await db.delete_subscription(subscription_id)  # ✅ Asinxron bazadan o‘chirish
        await callback.answer("✅ Kanal muvaffaqiyatli o‘chirildi!", show_alert=True)
    except Exception as e:
        await callback.answer(f"❌ Xatolik yuz berdi: {e}", show_alert=True)

    # Yangilangan ro‘yxatni qayta chiqarish
    new_keyboard = await get_delete_channel_keyboard()
    if new_keyboard:
        await callback.message.edit_text(
            "🗑 O'chirmoqchi bo'lgan kanalingizni tanlang:", reply_markup=new_keyboard
        )
    else:
        await callback.message.edit_text("✅ Barcha kanallar o‘chirildi!")


@router.message(AdminFilter(), F.text == "📊 Statistika")
async def show_statistics(message: Message):
    try:
        # Asosiy statistikalar
        total_users = await db.count_users()
        today_users = await db.count_users_by_date(datetime.now().date())

        # So'nggi 7 kunlik statistika
        weekly_stats = []
        for i in range(7):
            date = datetime.now().date() - pd.Timedelta(days=i)
            count = await db.count_users_by_date(date)
            if count > 0:
                weekly_stats.append(f"📅 {date.strftime('%d.%m.%Y')}: +{count}")

        stats = [
            "📊 Bot statistikasi\n",
            f"👥 Jami foydalanuvchilar: {total_users:,} ta",
            f"📅 Bugun qo'shilganlar: {today_users} ta\n",
            "📈 So'nggi 7 kunlik statistika:",
            *weekly_stats,
        ]

        await message.answer("\n".join(stats))

    except Exception as e:
        print(f"Error showing statistics: {e}")
        await message.answer("❌ Statistikani olishda xatolik yuz berdi")


@router.message(AdminFilter(), F.text == "📥 Users Excel")
async def get_users_excel(message: Message):
    await message.answer("📊 Excel fayl tayyorlanmoqda...")

    try:
        users = await db.get_all_users()
        if not users:
            await message.answer("❌ Foydalanuvchilar topilmadi")
            return

        # Ma'lumotlarni list of dict ko'rinishiga o'tkazamiz
        users_data = []
        for user in users:
            try:
                users_data.append(
                    {
                        "ID": user["id"] if "id" in user else "",
                        "Telegram ID": user["user_id"] if "user_id" in user else "",
                        "Username": user["username"] if "username" in user else "",
                        "To'liq ismi": user["full_name"] if "full_name" in user else "",
                        "Telefon raqami": (
                            user["phone_number"] if "phone_number" in user else ""
                        ),
                        "Ro'yxatdan o'tgan vaqti": (
                            user["created_at"] if "created_at" in user else ""
                        ),
                        "Oxirgi faolligi": (
                            user["last_active_at"] if "last_active_at" in user else ""
                        ),
                        "Holati": (
                            "Faol" if user.get("is_active", False) else "Faol emas"
                        ),
                        "Premium": "Ha" if user.get("is_premium", False) else "Yo'q",
                    }
                )
            except Exception as e:
                print(f"Error processing user data: {e}")
                continue

        df = pd.DataFrame(users_data)

        # Excel fayl nomi
        filename = f"users_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        filepath = f"data/files/{filename}"

        # Excel faylni yaratish
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Foydalanuvchilar", index=False)

            workbook = writer.book
            worksheet = writer.sheets["Foydalanuvchilar"]

            # Ustun kengliklarini moslash
            for idx, col in enumerate(df.columns):
                max_length = (
                    max(df[col].astype(str).apply(len).max(), len(str(col))) + 2
                )
                worksheet.column_dimensions[
                    worksheet.cell(1, idx + 1).column_letter
                ].width = max_length

            # Sarlavhalarni formatlash
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
                )

        # Faylni yuborish
        if os.path.exists(filepath):
            excel_file = FSInputFile(filepath)
            await message.answer_document(
                document=excel_file,
                caption=(
                    f"📊 Bot foydalanuvchilari ro'yxati:\n"
                    f"📅 Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                    f"👥 Jami: {len(users):,} ta foydalanuvchi"
                ),
            )
            try:
                os.remove(filepath)
            except FileNotFoundError:
                pass
        else:
            raise FileNotFoundError(f"Excel file not found at {filepath}")

    except Exception as e:
        print(f"Error creating Excel file: {e}")
        await message.answer("❌ Excel fayl yaratishda xatolik yuz berdi")


# 📌 📋 Kanallar ro‘yxati
@router.message(AdminFilter(), F.text == "📋 Kanallar ro'yxati")
async def get_channels(message: Message):
    buttons = await channels_button()
    await message.answer("Barcha kanallar:\n", reply_markup=buttons)


@router.message(AdminFilter(), F.text == "⬅️ Orqaga")
async def back_handler(message: Message, state: FSMContext):
    # FSM holatini tozalash
    await state.clear()

    # Admin paneldan chiqishni bildirish
    await message.answer("Siz admin  siz sizga admin panel korinib turadi")
